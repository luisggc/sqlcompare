from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from sqlcompare.cli import app
from sqlcompare.config import load_test_runs
from tests.cli_helpers import seed_duckdb, set_cli_env


def _create_diff(tmp_path: Path, monkeypatch) -> str:
    db_path = tmp_path / "sqlcompare.duckdb"
    seed_duckdb(db_path)
    config_dir = tmp_path / "config"
    set_cli_env(
        monkeypatch,
        config_dir,
        "duckdb_test",
        f"duckdb:///{db_path}",
    )
    runner = CliRunner()
    result = runner.invoke(
        app,
        [
            "table",
            "previous",
            "current",
            "id",
            "--connection",
            "duckdb_test",
        ],
    )
    assert result.exit_code == 0, result.output
    runs = load_test_runs()
    assert len(runs) == 1
    return next(iter(runs.keys()))


def test_analyze_diff_stats_and_list_columns(tmp_path, monkeypatch) -> None:
    diff_id = _create_diff(tmp_path, monkeypatch)
    runner = CliRunner()

    stats_result = runner.invoke(
        app,
        [
            "inspect",
            diff_id,
            "--stats",
            "--column",
            "name",
            "--limit",
            "5",
        ],
    )
    assert stats_result.exit_code == 0, stats_result.output
    assert "Statistics for diff ID" in stats_result.output

    list_result = runner.invoke(app, ["inspect", diff_id, "--list-columns"])
    assert list_result.exit_code == 0, list_result.output
    assert "Available columns" in list_result.output


def test_analyze_diff_save_modes_and_missing_filters(tmp_path, monkeypatch) -> None:
    diff_id = _create_diff(tmp_path, monkeypatch)
    runner = CliRunner()

    with runner.isolated_filesystem():
        save_summary = runner.invoke(
            app,
            [
                "inspect",
                diff_id,
                "--column",
                "value",
                "--save",
                "summary",
                "--file-path",
                "reports/inspect_summary",
            ],
        )
        assert save_summary.exit_code == 0, save_summary.output
        assert "Report saved to:" in save_summary.output
        summary_path = Path.cwd() / "reports" / "inspect_summary.xlsx"
        assert summary_path.exists()

        summary_wb = load_workbook(summary_path)
        assert "Overview" in summary_wb.sheetnames
        assert "SQL Reference" in summary_wb.sheetnames
        assert any(name.lower() == "value" for name in summary_wb.sheetnames)

        save_complete = runner.invoke(
            app,
            [
                "inspect",
                diff_id,
                "--save",
                "complete",
            ],
        )
        assert save_complete.exit_code == 0, save_complete.output
        assert "Report saved to:" in save_complete.output
        saved = list(Path.cwd().glob("inspect_report_*.xlsx"))
        assert saved

    missing_current = runner.invoke(app, ["inspect", diff_id, "--missing-current"])
    assert missing_current.exit_code == 0, missing_current.output
    assert "Loaded diff data" in missing_current.output

    missing_previous = runner.invoke(
        app, ["inspect", diff_id, "--missing-previous"]
    )
    assert missing_previous.exit_code == 0, missing_previous.output
    assert "Loaded diff data" in missing_previous.output


def test_analyze_diff_save_mode_validation(tmp_path, monkeypatch) -> None:
    diff_id = _create_diff(tmp_path, monkeypatch)
    runner = CliRunner()

    invalid_mode = runner.invoke(app, ["inspect", diff_id, "--save", "csv"])
    assert invalid_mode.exit_code != 0
    assert "Invalid save mode" in invalid_mode.output

    incompatible = runner.invoke(
        app, ["inspect", diff_id, "--save", "summary", "--stats"]
    )
    assert incompatible.exit_code != 0
    assert "Report export (--save summary/complete)" in incompatible.output
