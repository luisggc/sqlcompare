import glob
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from sqlcompare.config import get_tests_folder, load_test_runs
from sqlcompare.log import log
from sqlcompare.utils.format import format_table

SUMMARY_TAB_LIMIT = 200
XLSX_MAX_ROWS = 1_048_576
XLSX_MAX_COLUMNS = 16_384


def find_diff_file(diff_id: str) -> str | None:
    tests_folder = get_tests_folder()
    pattern = f"{tests_folder}/*/diffs/{diff_id}.pkl"
    matches = glob.glob(pattern)
    if matches:
        return matches[0]
    pattern = f"{tests_folder}/*/diffs/*{diff_id}*.pkl"
    matches = glob.glob(pattern)
    if matches:
        if len(matches) == 1:
            return matches[0]
        log.warning(f"⚠️  Multiple files match '{diff_id}':")
        for match in matches:
            file_id = os.path.basename(match).replace(".pkl", "")
            log.info(f"   {file_id}")
        log.info("💡 Please use a more specific ID")
    return None


def find_diff_run(diff_id: str):
    runs = load_test_runs()
    if diff_id in runs:
        run = runs[diff_id]
        run["id"] = diff_id
        return run
    for key, value in runs.items():
        if diff_id in key:
            value["id"] = key
            return value
    return None


def _display(
    columns: list[str],
    rows: list[tuple],
    column: str | None,
    limit: int,
    diff_id: str,
    *,
    is_stats: bool = False,
    list_columns: bool = False,
    total_differences: int | None = None,
):
    if not columns:
        log.info("No results returned.")
        return

    column_idx = _column_index(columns, "Column")
    column_name_idx = _column_index(columns, "COLUMN_NAME")
    if column_idx is None and column_name_idx is not None:
        column_idx = column_name_idx

    if is_stats:
        log.info(f"\U0001f4ca Statistics for diff ID: {diff_id}")
        filtered_rows = rows
        if column and column_idx is not None:
            column_upper = column.upper()
            filtered_rows = [
                row for row in rows if column_upper in str(row[column_idx]).upper()
            ]
            if not filtered_rows:
                log.info(f"❌ No statistics found for column containing '{column}'")
                return
        log.info(format_table(columns, filtered_rows[:limit]))
        return

    effective_total = total_differences if total_differences is not None else len(rows)
    log.info(f"📂 Loaded diff data: {effective_total} total differences")
    if list_columns:
        if column_idx is None:
            log.warning("⚠️  Column metadata not available.")
            return
        counts: dict[str, int] = {}
        for row in rows:
            col_name = str(row[column_idx])
            counts[col_name] = counts.get(col_name, 0) + 1
        log.info(f"📋 Available columns ({len(counts)}):")
        for col_name in sorted(counts.keys()):
            log.info(f"   {col_name}: {counts[col_name]} differences")
        return

    filtered_rows = rows
    filtered_total = effective_total
    if column and column_idx is not None:
        column_upper = column.upper()
        filtered_rows = [
            row for row in rows if str(row[column_idx]).upper() == column_upper
        ]
        if total_differences is None:
            filtered_total = len(filtered_rows)
        if not filtered_rows:
            log.warning(f"⚠️  No differences found for column '{column}'")
            return
        log.info(f"🔍 Filtered to {filtered_total} differences for column '{column}'")

    display_rows = filtered_rows[:limit]
    log.info(f"📊 Showing {len(display_rows)} of {filtered_total} differences:")
    log.info(format_table(columns, display_rows))
    if filtered_total > limit:
        log.info(f"💡 Use --limit {filtered_total} to see all results")


def _column_index(columns: list[str], name: str) -> int | None:
    name_upper = name.upper()
    for idx, col in enumerate(columns):
        if col.upper() == name_upper:
            return idx
    return None


def resolve_report_path(
    diff_id: str, save_mode: str, file_path: str | None = None
) -> Path:
    if file_path:
        path = Path(file_path).expanduser()
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = Path.cwd() / f"inspect_report_{diff_id}_{save_mode}_{timestamp}.xlsx"

    if path.suffix == "":
        path = path.with_suffix(".xlsx")
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Inspect report file must use the .xlsx extension.")
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def write_inspect_report_xlsx(
    output_path: Path,
    overview_rows: list[tuple[Any, Any, Any]],
    column_diffs: dict[str, tuple[list[str], list[tuple[Any, ...]]]],
    sql_reference_rows: list[tuple[str, str, str]],
) -> dict[str, int]:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for inspect report export. Install sqlcompare with openpyxl available."
        ) from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    truncated_rows = 0

    overview_sheet = workbook.create_sheet(title="Overview")
    overview_sheet.append(["Section", "Name", "Value"])
    for row in overview_rows:
        overview_sheet.append(list(row[:3]))

    used_sheet_names = {"Overview", "SQL Reference"}
    for column_name, (headers, rows) in column_diffs.items():
        sheet_name = _unique_sheet_name(column_name, used_sheet_names)
        used_sheet_names.add(sheet_name)
        sheet = workbook.create_sheet(title=sheet_name)

        clipped_headers = [str(h) for h in headers[:XLSX_MAX_COLUMNS]]
        sheet.append(clipped_headers)

        max_data_rows = XLSX_MAX_ROWS - 1
        clipped_rows = rows[:max_data_rows]
        for row in clipped_rows:
            sheet.append(list(row[:XLSX_MAX_COLUMNS]))

        if len(rows) > max_data_rows:
            truncated_rows += len(rows) - max_data_rows

    sql_sheet = workbook.create_sheet(title="SQL Reference")
    sql_sheet.append(["Category", "Name", "SQL"])
    for row in sql_reference_rows:
        sql_sheet.append(list(row[:3]))

    workbook.save(output_path)
    return {"truncated_rows": truncated_rows}


def _unique_sheet_name(raw_name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]\*\:/\\\?]", "_", raw_name).strip() or "Sheet"
    cleaned = cleaned[:31]
    if cleaned not in used:
        return cleaned

    suffix = 1
    while True:
        candidate_suffix = f"_{suffix}"
        base = cleaned[: 31 - len(candidate_suffix)]
        candidate = f"{base}{candidate_suffix}"
        if candidate not in used:
            return candidate
        suffix += 1


def list_available_diffs():
    tests_folder = get_tests_folder()
    pattern = f"{tests_folder}/*/diffs/*.pkl"
    matches = glob.glob(pattern)
    diff_ids = []
    for match in matches:
        diff_id = os.path.basename(match).replace(".pkl", "")
        if (
            "_stats" in diff_id
            or "_in_current_only" in diff_id
            or "_in_previous_only" in diff_id
        ):
            continue
        test_name = match.split("/")[-3]
        file_size_mb = os.path.getsize(match) / (1024 * 1024)
        diff_ids.append((diff_id, test_name, file_size_mb))
    runs = load_test_runs()
    for run_id, run in runs.items():
        diff_ids.append((run_id, run.get("conn", "db"), 0))
    diff_ids.sort(reverse=True)
    if not diff_ids:
        log.info("   No saved diff data found.")
        return
    log.info("   Recent diff IDs:")
    for diff_id, test_name, size_mb in diff_ids[:10]:
        size_display = f"{size_mb:.1f}MB" if size_mb else "db"
        log.info(f"   {diff_id} (test: {test_name}, {size_display})")
    if len(diff_ids) > 10:
        log.info(f"   ... and {len(diff_ids) - 10} more")
